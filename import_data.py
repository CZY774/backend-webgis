import sys
import os

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from database import SessionLocal, engine
from models import Base, Fasilitas, UMKM, Wisata, SDA, RW, Kependudukan
from geoalchemy2.shape import from_shape
from shapely.geometry import shape
import openpyxl
import json

# Create tables
Base.metadata.create_all(bind=engine)

db = SessionLocal()

print("Importing data...")

# Import Fasilitas
print("\n1. Importing Fasilitas...")
wb = openpyxl.load_workbook("../Fasilitas.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:  # Skip header
    fasilitas = Fasilitas(latitude=row[0], longitude=row[1], nama=row[2], jenis=row[3])
    db.add(fasilitas)
db.commit()
print(f"✓ Imported {len(rows) - 1} fasilitas")

# Import UMKM
print("\n2. Importing UMKM...")
wb = openpyxl.load_workbook("../UMKM.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:  # Skip header
    umkm = UMKM(latitude=row[1], longitude=row[2], nama=row[3], jenis=row[4])
    db.add(umkm)
db.commit()
print(f"✓ Imported {len(rows) - 1} UMKM")

# Import Wisata
print("\n3. Importing Wisata...")
wb = openpyxl.load_workbook("../Wisata.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:  # Skip header
    wisata = Wisata(
        latitude=row[0],
        longitude=row[1],
        nama=row[2],
        jenis=row[3],
        deskripsi=row[4] if len(row) > 4 else None,
    )
    db.add(wisata)
db.commit()
print(f"✓ Imported {len(rows) - 1} wisata")

# Import SDA (Sawah, Kebun, Ladang, Pemukiman)
print("\n4. Importing SDA...")
sda_files = ["Sawah", "Kebun", "Ladang", "Pemukiman"]
total_sda = 0

for jenis in sda_files:
    with open(f"../{jenis}.geojson", "r") as f:
        data = json.load(f)
        for feature in data["features"]:
            geom = shape(feature["geometry"])
            # Calculate area in square degrees, then convert to hectares (approximate)
            # For more accurate calculation, would need to use pyproj
            area_sq_deg = geom.area
            luas_ha = area_sq_deg * 12321  # Rough conversion at this latitude

            sda = SDA(
                polygon=from_shape(geom, srid=4326),
                jenis_lahan=jenis,
                luas_ha=round(luas_ha, 4),
            )
            db.add(sda)
            total_sda += 1
db.commit()
print(f"✓ Imported {total_sda} SDA polygons")

# Import RW and Kependudukan
print("\n5. Importing RW and Kependudukan...")

# Check if RW data already exists
existing_rw = db.query(RW).count()
if existing_rw > 0:
    print(f"✓ RW data already exists ({existing_rw} records), skipping...")
else:
    with open("../BatasRW.geojson", "r") as f:
        data = json.load(f)
        for feature in data["features"]:
            rw_id = feature["properties"]["id"]
            geom = shape(feature["geometry"])

            rw = RW(nomor_rw=rw_id, polygon=from_shape(geom, srid=4326))
            db.add(rw)
            db.flush()  # Get the id_rw

    db.commit()
    print(f"✓ Imported {len(data['features'])} RW polygons")

# Import Kependudukan data
print("\n6. Importing Kependudukan data...")

# Check if kependudukan data already exists
existing_kependudukan = db.query(Kependudukan).count()
if existing_kependudukan > 0:
    print(
        f"✓ Kependudukan data already exists ({existing_kependudukan} records), skipping..."
    )
else:
    wb = openpyxl.load_workbook("../Kependudukan.xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

# Parse the data structure
rw_data = {}
for i, row in enumerate(rows):
    if i == 0:  # Header row
        continue

    kategori = row[0]
    subkategori = row[1]

    if kategori == "Jumlah Warga":
        for rw_num in range(1, 7):
            if rw_num not in rw_data:
                rw_data[rw_num] = {}
            rw_data[rw_num]["jumlah_warga"] = row[rw_num + 1]

    elif kategori == "Jenis Kelamin":
        for rw_num in range(1, 7):
            if subkategori == "Laki-laki":
                rw_data[rw_num]["laki_laki"] = row[rw_num + 1]
            elif subkategori == "Perempuan":
                rw_data[rw_num]["perempuan"] = row[rw_num + 1]

    elif kategori == "Umur":
        for rw_num in range(1, 7):
            if "Anak-anak" in str(subkategori):
                rw_data[rw_num]["anak_anak"] = row[rw_num + 1]
            elif "Produktif" in str(subkategori):
                rw_data[rw_num]["produktif"] = row[rw_num + 1]
            elif "Lansia" in str(subkategori):
                rw_data[rw_num]["lansia"] = row[rw_num + 1]

    elif kategori == "Pendidikan":
        for rw_num in range(1, 7):
            if "Tidak/Belum" in str(subkategori):
                rw_data[rw_num]["tidak_sekolah"] = row[rw_num + 1]
            elif "Tidak Tamat" in str(subkategori):
                rw_data[rw_num]["tidak_tamat_sd"] = row[rw_num + 1]
            elif "Tamat SD" in str(subkategori):
                rw_data[rw_num]["tamat_sd"] = row[rw_num + 1]
            elif "SLTP" in str(subkategori):
                rw_data[rw_num]["sltp"] = row[rw_num + 1]
            elif "SLTA" in str(subkategori):
                rw_data[rw_num]["slta"] = row[rw_num + 1]
            elif "Diploma" in str(subkategori):
                rw_data[rw_num]["diploma_s1"] = row[rw_num + 1]

    elif kategori == "Pekerjaan":
        for rw_num in range(1, 7):
            if "Belum/Tidak" in str(subkategori):
                rw_data[rw_num]["belum_bekerja"] = row[rw_num + 1]
            elif "Pelajar" in str(subkategori):
                rw_data[rw_num]["pelajar"] = row[rw_num + 1]
            elif "Mengurus" in str(subkategori):
                rw_data[rw_num]["mengurus_rt"] = row[rw_num + 1]
            elif "Wiraswasta" in str(subkategori):
                rw_data[rw_num]["wiraswasta"] = row[rw_num + 1]
            elif "Petani" in str(subkategori):
                rw_data[rw_num]["petani"] = row[rw_num + 1]
            elif "Lainnya" in str(subkategori):
                rw_data[rw_num]["lainnya"] = row[rw_num + 1]

    # Insert kependudukan data
    for rw_num, data in rw_data.items():
        rw = db.query(RW).filter(RW.nomor_rw == rw_num).first()
        if rw:
            kependudukan = Kependudukan(
                id_rw=rw.id_rw,
                jumlah_warga=data.get("jumlah_warga", 0),
                laki_laki=data.get("laki_laki", 0),
                perempuan=data.get("perempuan", 0),
                anak_anak=data.get("anak_anak", 0),
                produktif=data.get("produktif", 0),
                lansia=data.get("lansia", 0),
                tidak_sekolah=data.get("tidak_sekolah", 0),
                tidak_tamat_sd=data.get("tidak_tamat_sd", 0),
                tamat_sd=data.get("tamat_sd", 0),
                sltp=data.get("sltp", 0),
                slta=data.get("slta", 0),
                diploma_s1=data.get("diploma_s1", 0),
                belum_bekerja=data.get("belum_bekerja", 0),
                pelajar=data.get("pelajar", 0),
                mengurus_rt=data.get("mengurus_rt", 0),
                wiraswasta=data.get("wiraswasta", 0),
                petani=data.get("petani", 0),
                lainnya=data.get("lainnya", 0),
            )
            db.add(kependudukan)

    db.commit()
    print(f"✓ Imported kependudukan data for 6 RW")

# Create default admin
print("\n7. Creating default admin...")
import bcrypt

from models import Admin

existing_admin = db.query(Admin).filter(Admin.username == "admin").first()
if not existing_admin:
    hashed = bcrypt.hashpw("admin123".encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    admin = Admin(username="admin", password=hashed)
    db.add(admin)
    db.commit()
    print("✓ Created admin user (username: admin, password: admin123)")
else:
    print("✓ Admin user already exists")

db.close()
print("\n✅ Data import completed successfully!")
