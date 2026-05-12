import sys
import os

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Use the new models file
import models_rev as models
from database import SessionLocal, engine
from geoalchemy2.shape import from_shape
from shapely.geometry import shape
from shapely import wkt
import openpyxl
import json

# Helper function to force 2D geometry
def force_2d(geom):
    """Strip Z dimension from geometry"""
    return wkt.loads(wkt.dumps(geom, output_dimension=2))

# Create tables
models.Base.metadata.create_all(bind=engine)

db = SessionLocal()

print("Importing data from revised files...")

# Import Fasilitas
print("\n1. Importing Fasilitas...")
wb = openpyxl.load_workbook("../../10. Data Web/2. Fasilitas/Fasilitas_rev.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:  # Skip header
    fasilitas = models.Fasilitas(latitude=row[0], longitude=row[1], nama=row[2], jenis=row[3])
    db.add(fasilitas)
db.commit()
print(f"✓ Imported {len(rows) - 1} fasilitas")

# Import UMKM
print("\n2. Importing UMKM...")
wb = openpyxl.load_workbook("../../10. Data Web/3. UMKM/UMKM_rev.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:  # Skip header
    umkm = models.UMKM(latitude=row[1], longitude=row[2], nama=row[3], jenis=row[4])
    db.add(umkm)
db.commit()
print(f"✓ Imported {len(rows) - 1} UMKM")

# Import Wisata
print("\n3. Importing Wisata...")
wb = openpyxl.load_workbook("../../10. Data Web/1. Wisata/Wisata_rev.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:  # Skip header
    wisata = models.Wisata(
        latitude=row[0],
        longitude=row[1],
        nama=row[2],
        jenis=row[3],
        deskripsi=row[4],
    )
    db.add(wisata)
db.commit()
print(f"✓ Imported {len(rows) - 1} wisata")

# Import SDA (Land Use) with nomor-to-jenis mapping
print("\n4. Importing SDA (Land Use)...")

# Load the mapping from ini warnanya.xlsx
wb = openpyxl.load_workbook("../../10. Data Web/4. Penggunaan Lahan/ini warnanya.xlsx")
ws = wb.active
nomor_to_jenis = {}
for row in ws.iter_rows(values_only=True):
    if row[0] and row[0] != "Nomor":  # Skip header
        nomor_to_jenis[str(row[0])] = row[1]

# Import lahan.geojson with mapping
with open("../../10. Data Web/4. Penggunaan Lahan/lahan.geojson", "r") as f:
    data = json.load(f)
    total_sda = 0
    skipped = 0
    
    for feature in data["features"]:
        nomor = feature["properties"].get("nomor")
        jenis_lahan = nomor_to_jenis.get(str(nomor))
        
        if not jenis_lahan:
            skipped += 1
            continue
            
        geom = force_2d(shape(feature["geometry"]))
        
        area_sq_deg = geom.area
        luas_ha = area_sq_deg * 12321

        sda = models.SDA(
            polygon=from_shape(geom, srid=4326),
            jenis_lahan=jenis_lahan,
            luas_ha=round(luas_ha, 4),
        )
        db.add(sda)
        total_sda += 1
        
db.commit()
print(f"✓ Imported {total_sda} SDA polygons")
if skipped > 0:
    print(f"  (Skipped {skipped} features with unmapped nomor)")

# Import Jalan (Roads)
print("\n5. Importing Jalan (Roads)...")
with open("../../10. Data Web/4. Penggunaan Lahan/jalan.geojson", "r") as f:
    data = json.load(f)
    for feature in data["features"]:
        props = feature["properties"]
        geom = force_2d(shape(feature["geometry"]))
        
        jalan = models.Jalan(
            linestring=from_shape(geom, srid=4326),
            nama_jalan=props.get("nama_jalan"),
            jenis=props.get("jenis"),
            permukaan=props.get("permukaan"),
            lebar_m=props.get("lebar(m)"),
        )
        db.add(jalan)
db.commit()
print(f"✓ Imported {len(data['features'])} jalan")

# Import Sungai (Rivers)
print("\n6. Importing Sungai (Rivers)...")
with open("../../10. Data Web/4. Penggunaan Lahan/sungai.geojson", "r") as f:
    data = json.load(f)
    for feature in data["features"]:
        geom = force_2d(shape(feature["geometry"]))
        sungai = models.Sungai(linestring=from_shape(geom, srid=4326))
        db.add(sungai)
db.commit()
print(f"✓ Imported {len(data['features'])} sungai")

# Import RW
print("\n7. Importing RW...")
existing_rw = db.query(models.RW).count()
if existing_rw > 0:
    print(f"✓ RW data already exists ({existing_rw} records), skipping...")
else:
    with open("../../10. Data Web/5. Kependudukan/Batas_RW.geojson", "r") as f:
        data = json.load(f)
        for feature in data["features"]:
            rw_num = feature["properties"]["rw"]  # Use 'rw' property, not 'id'
            geom = force_2d(shape(feature["geometry"]))
            rw = models.RW(nomor_rw=rw_num, polygon=from_shape(geom, srid=4326))
            db.add(rw)
            db.flush()
    db.commit()
    print(f"✓ Imported {len(data['features'])} RW polygons")

# Import RT
print("\n8. Importing RT...")
with open("../../10. Data Web/5. Kependudukan/Batas_RT.geojson", "r") as f:
    data = json.load(f)
    for feature in data["features"]:
        props = feature["properties"]
        rt_id = props["id"]
        rw_num = props["rw"]
        
        # Find the RW record
        rw = db.query(models.RW).filter(models.RW.nomor_rw == rw_num).first()
        if rw:
            geom = force_2d(shape(feature["geometry"]))
            rt = models.RT(
                nomor_rt=rt_id,
                id_rw=rw.id_rw,
                polygon=from_shape(geom, srid=4326)
            )
            db.add(rt)
db.commit()
print(f"✓ Imported {len(data['features'])} RT polygons")

# Import Desa (Village Boundary)
print("\n9. Importing Desa (Village Boundary)...")
with open("../../10. Data Web/5. Kependudukan/Batas_Desa.geojson", "r") as f:
    data = json.load(f)
    for feature in data["features"]:
        props = feature["properties"]
        geom = force_2d(shape(feature["geometry"]))
        
        desa = models.Desa(
            polygon=from_shape(geom, srid=4326),
            nama_desa=props.get("WADMKD", "Prawoto"),
            kecamatan=props.get("WADMKC"),
            kabupaten=props.get("WADMKK"),
            provinsi=props.get("WADMPR"),
            luas_ha=props.get("LUASWH"),
        )
        db.add(desa)
db.commit()
print(f"✓ Imported {len(data['features'])} desa boundary")

# Import Kependudukan data
print("\n10. Importing Kependudukan data...")
existing_kependudukan = db.query(models.Kependudukan).count()
if existing_kependudukan > 0:
    print(f"✓ Kependudukan data already exists ({existing_kependudukan} records), skipping...")
else:
    wb = openpyxl.load_workbook("../../10. Data Web/5. Kependudukan/Kependudukan_rev.xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    rw_data = {}
    for i, row in enumerate(rows):
        if i == 0:
            continue

        kategori = row[0]
        subkategori = row[1]

        if kategori == "Jumlah Warga":
            for rw_num in range(1, 7):
                if rw_num not in rw_data:
                    rw_data[rw_num] = {}
                rw_data[rw_num]["jumlah_warga"] = row[rw_num + 1]

        elif kategori == "Jenis Kelamin":
            for rw_num in range(1, 7):
                if subkategori == "Laki-laki":
                    rw_data[rw_num]["laki_laki"] = row[rw_num + 1]
                elif subkategori == "Perempuan":
                    rw_data[rw_num]["perempuan"] = row[rw_num + 1]

        elif kategori == "Umur":
            for rw_num in range(1, 7):
                if "Anak-anak" in str(subkategori):
                    rw_data[rw_num]["anak_anak"] = row[rw_num + 1]
                elif "Produktif" in str(subkategori):
                    rw_data[rw_num]["produktif"] = row[rw_num + 1]
                elif "Lansia" in str(subkategori):
                    rw_data[rw_num]["lansia"] = row[rw_num + 1]

        elif kategori == "Pendidikan":
            for rw_num in range(1, 7):
                if "Tidak/Belum" in str(subkategori):
                    rw_data[rw_num]["tidak_sekolah"] = row[rw_num + 1]
                elif "Tidak Tamat" in str(subkategori):
                    rw_data[rw_num]["tidak_tamat_sd"] = row[rw_num + 1]
                elif "Tamat SD" in str(subkategori):
                    rw_data[rw_num]["tamat_sd"] = row[rw_num + 1]
                elif "SLTP" in str(subkategori):
                    rw_data[rw_num]["sltp"] = row[rw_num + 1]
                elif "SLTA" in str(subkategori):
                    rw_data[rw_num]["slta"] = row[rw_num + 1]
                elif "Diploma" in str(subkategori):
                    rw_data[rw_num]["diploma_s1"] = row[rw_num + 1]

        elif kategori == "Pekerjaan":
            for rw_num in range(1, 7):
                if "Belum/Tidak" in str(subkategori):
                    rw_data[rw_num]["belum_bekerja"] = row[rw_num + 1]
                elif "Pelajar" in str(subkategori):
                    rw_data[rw_num]["pelajar"] = row[rw_num + 1]
                elif "Mengurus" in str(subkategori):
                    rw_data[rw_num]["mengurus_rt"] = row[rw_num + 1]
                elif "Wiraswasta" in str(subkategori):
                    rw_data[rw_num]["wiraswasta"] = row[rw_num + 1]
                elif "Petani" in str(subkategori):
                    rw_data[rw_num]["petani"] = row[rw_num + 1]
                elif "Lainnya" in str(subkategori):
                    rw_data[rw_num]["lainnya"] = row[rw_num + 1]

    for rw_num, data in rw_data.items():
        rw = db.query(models.RW).filter(models.RW.nomor_rw == rw_num).first()
        if rw:
            kependudukan = models.Kependudukan(
                id_rw=rw.id_rw,
                jumlah_warga=data.get("jumlah_warga", 0),
                laki_laki=data.get("laki_laki", 0),
                perempuan=data.get("perempuan", 0),
                anak_anak=data.get("anak_anak", 0),
                produktif=data.get("produktif", 0),
                lansia=data.get("lansia", 0),
                tidak_sekolah=data.get("tidak_sekolah", 0),
                tidak_tamat_sd=data.get("tidak_tamat_sd", 0),
                tamat_sd=data.get("tamat_sd", 0),
                sltp=data.get("sltp", 0),
                slta=data.get("slta", 0),
                diploma_s1=data.get("diploma_s1", 0),
                belum_bekerja=data.get("belum_bekerja", 0),
                pelajar=data.get("pelajar", 0),
                mengurus_rt=data.get("mengurus_rt", 0),
                wiraswasta=data.get("wiraswasta", 0),
                petani=data.get("petani", 0),
                lainnya=data.get("lainnya", 0),
            )
            db.add(kependudukan)

    db.commit()
    print(f"✓ Imported kependudukan data for 6 RW")

# Create default admin
print("\n11. Creating default admin...")
import bcrypt

existing_admin = db.query(models.Admin).filter(models.Admin.username == "admin").first()
if not existing_admin:
    hashed = bcrypt.hashpw("admin123".encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    admin = models.Admin(username="admin", password=hashed)
    db.add(admin)
    db.commit()
    print("✓ Created admin user (username: admin, password: admin123)")
else:
    print("✓ Admin user already exists")

print("\n✅ Data import completed successfully!")
print("\nSummary:")
print(f"  - Fasilitas: {db.query(models.Fasilitas).count()}")
print(f"  - UMKM: {db.query(models.UMKM).count()}")
print(f"  - Wisata: {db.query(models.Wisata).count()}")
print(f"  - SDA (Land Use): {db.query(models.SDA).count()}")
print(f"  - Jalan: {db.query(models.Jalan).count()}")
print(f"  - Sungai: {db.query(models.Sungai).count()}")
print(f"  - RW: {db.query(models.RW).count()}")
print(f"  - RT: {db.query(models.RT).count()}")
print(f"  - Desa: {db.query(models.Desa).count()}")
print(f"  - Kependudukan: {db.query(models.Kependudukan).count()}")

db.close()
